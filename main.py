import math
import openpyxl as xl
from datetime import datetime
import data_correction as dc
import production_code as pc


brewery_dictionary = {
		"Creston BC Canada": "Creston",
		"Edmonton, Alberta, Canada (Labatt)": "Edmonton",
		"Halifax, Nova Scotia, Canada (Labatt)": "Halifax",
		"London, Ontario, Canada (Labatt)": "London",
		"Montreal, Quebec, Canada (Labatt)": "Montreal",
		"St Johns, Newfoundland, Canada (Labatt)": "St. John's",
		"Turning Point (Stanley Park)": "Turning Point"
	}

plant_list = ["Turning Point", "Mill Street", "Archibald", "Creston", "London", "St. John's", "Halifax", "Montreal", "Edmonton", "Unconfirmed"]

class Complaint():

	def __init__(self, sheet, r, one_brewery_brands, one_country_brands):

		self.customer_code = self.getCustomerCode(sheet, r)
		self.brand = self.getBrand(sheet, r)
		self.brewery = self.getBrewery(sheet, r, one_brewery_brands)
		self.case_number = self.getCaseNumber(sheet, r)
		self.complaint_date = self.getComplaintDate(sheet, r)
		self.country = self.getCountry(sheet, r, one_country_brands)
		self.incident_date = self.getIncidentDate(sheet, r)
		self.production_date = self.getProductionDate(sheet, r)
		self.age = self.getAge()

	def getCustomerCode(self, sheet, r):
		return pc.ProductionCode(sheet.cell(row = r, column = 12).value)
	
	def getBrand(self, sheet, r):
		return sheet.cell(row = r, column = 4).value

	def getBrewery(self, sheet, r, one_brewery_brands):
		
		if self.brand in one_brewery_brands.keys():
			return one_brewery_brands[self.brand]
		elif self.customer_code.validateProductionCode():
			return self.customer_code.brewery
		else:
			facility = sheet.cell(row = r, column = 6).value
			if facility in brewery_dictionary.keys():
				return brewery_dictionary[facility]
			elif facility is None:
				return "Unconfirmed"
			else:
				return facility

	def getCaseNumber(self, sheet, r):
		return sheet.cell(row = r, column = 2).value

	def getComplaintDate(self, sheet, r):
		complaint_date = sheet.cell(row = r, column = 3).value
		if complaint_date == "":
			return complaint_date
		else:
			return convertExcelDate(complaint_date)

	def getCountry(self, sheet, r, one_country_brands):
		
		data_country = sheet.cell(row = r, column = 14).value

		if self.customer_code.validateProductionCode():
			return "CANADA"
		elif self.brand in one_country_brands.keys():
			return one_country_brands[self.brand]
		elif data_country == "UNITED STATES":
			return "USA"
		elif self.brewery not in plant_list:
			return None
		else:
			return data_country

		

	def getIncidentDate(self, sheet, r):
		incident_date = sheet.cell(row = r, column = 31).value
		if incident_date is None:
			return incident_date
		else:
			return convertExcelDate(incident_date)

	def getProductionDate(self, sheet, r):
		if self.customer_code.validateProductionCode():
			return self.customer_code.date
		else:
			production_date = sheet.cell(row = r, column = 29).value
			if production_date is None:
				return production_date
			else:
				return convertExcelDate(production_date)

	def getAge(self):
		if self.production_date is None:
			return None
		else:
			if self.incident_date is not None:
				return (self.incident_date - self.production_date).days
			else:
				return (self.complaint_date - self.production_date).days

	def isOverage(self):
		if self.age is None:
			return "UNKNOWN"
		elif self.age <= 180:
			return False
		else:
			return True


def headingPrinter(sheet):
	sheet.cell(row = 1, column = 1).value = "Case Number"
	sheet.cell(row = 1, column = 2).value = "Brand"
	sheet.cell(row = 1, column = 3).value = "Brewery"
	sheet.cell(row = 1, column = 4).value = "Country"
	sheet.cell(row = 1, column = 5).value = "Production Date"
	sheet.cell(row = 1, column = 6).value = "Complaint Date"
	sheet.cell(row = 1, column = 7).value = "Incident Date"
	sheet.cell(row = 1, column = 8).value = "Age"
	sheet.cell(row = 1, column = 9).value = "Overage?"

def complaintPrinter(complaint_data, sheet, r):

	sheet.cell(row = r, column = 1).value = complaint_data.case_number
	sheet.cell(row = r, column = 2).value = complaint_data.brand
	sheet.cell(row = r, column = 3).value = complaint_data.brewery
	sheet.cell(row = r, column = 4).value = complaint_data.country
	sheet.cell(row = r, column = 5).value = complaint_data.production_date
	sheet.cell(row = r, column = 6).value = complaint_data.complaint_date
	sheet.cell(row = r, column = 7).value = complaint_data.incident_date
	sheet.cell(row = r, column = 8).value = complaint_data.age
	sheet.cell(row = r, column = 9).value = complaint_data.isOverage()

def singleBreweryBrands(sheet):
	#[brand]: [facility]
	single_brewery_brands = {}
	for i in range(3, sheet.max_row + 1):
		single_brewery_brands[sheet.cell(column = 1, row = i).value] = sheet.cell(column = 2, row = i).value

	return single_brewery_brands

def singleCountryBrands(sheet):
	single_country_brands = {}
	i = 3
	while sheet.cell(column = 4, row = i).value:
		single_country_brands[sheet.cell(column = 4, row = i).value] = sheet.cell(column = 5, row = i).value
		i += 1

	return single_country_brands

def convertExcelDate(excel_date):
	if isinstance(excel_date, datetime):
		return excel_date
	else:
		return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + excel_date - 2)

def main():

	wb = xl.load_workbook("data.xlsx")
	ws = wb["Data"]
	ws_setup = wb["Setup"]
	ws_destiny = wb["Polished Data"]
	ws_corrections = wb["Manual Corrections"]
	one_brewery_brands_dic = singleBreweryBrands(ws_setup)
	one_country_brands_dic = singleCountryBrands(ws_setup)

	dc.correction(ws_corrections, ws)

	for i in range(4, ws.max_row+1):
		print(i)
		current_complaint = Complaint(ws, i, one_brewery_brands_dic, one_country_brands_dic)
		headingPrinter(ws_destiny)
		complaintPrinter(current_complaint, ws_destiny, i-2)

	wb.save("data.xlsx")

main()
