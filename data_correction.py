import openpyxl as xl

def correction(input_sh, output_sh):

	for r in range(2, output_sh.max_row+1):
		case_num = output_sh.cell(row = r, column = 2).value
		output_row = findRow(case_num, input_sh)
		for c in range(1, output_sh.max_column + 1):
			output_sh.cell(row = output_row, column = c).value = input_sh.cell(row = r, column = c).value
		print("{} corrections performed.".format(r - 1))

def findRow(case_num, sheet):
	for r in range(4, sheet.max_row + 1):
		if sheet.cell(row = r, column = 2).value == case_num:
			return r